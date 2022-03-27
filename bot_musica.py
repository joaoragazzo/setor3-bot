import asyncio
from asyncio.tasks import wait_for
import functools
import itertools
from logging import disable, exception
import math
import random
from re import M
import time
from typing import Optional, final
from discord import channel, colour, reaction
from discord.ext.commands.core import command
from datetime import date, datetime
from discord import FFmpegPCMAudio
from PIL import ImageGrab
import pandas as pd
import openpyxl 
from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup
import pygetwindow
import pyautogui
import PIL
from PIL import Image
import os
import subprocess
import pynput
from pynput import keyboard
from pynput.mouse import Button, Controllerm
from pynput.keyboard import Key, Controllerk
import requests 
from requests import *
import discord
from discord import activity
from discord.ext.commands.bot import Bot
from discord.ext import commands
from discord.flags import Intents
import youtube_dl
from async_timeout import timeout
from discord.ext import commands
import cv2
import urllib
import numpy as np
import sys
from threading import Thread

# Silence useless bug reports messages
youtube_dl.utils.bug_reports_message = lambda: ''


class VoiceError(Exception):
    pass

class YTDLError(Exception):
    pass

class YTDLSource(discord.PCMVolumeTransformer):
    YTDL_OPTIONS = {
        'format': 'bestaudio/best',
        'extractaudio': True,
        'audioformat': 'mp3',
        'outtmpl': '%(extractor)s-%(id)s-%(title)s.%(ext)s',
        'restrictfilenames': True,
        'noplaylist': True,
        'nocheckcertificate': True,
        'ignoreerrors': False,
        'logtostderr': False,
        'quiet': True,
        'no_warnings': True,
        'default_search': 'auto',
        'source_address': '0.0.0.0',
    }

    FFMPEG_OPTIONS = {
        'before_options': '-reconnect 1 -reconnect_streamed 1 -reconnect_delay_max 5',
        'options': '-vn',
    }

    ytdl = youtube_dl.YoutubeDL(YTDL_OPTIONS)

    def __init__(self, ctx: commands.Context, source: discord.FFmpegPCMAudio, *, data: dict, volume: float = 0.5):
        super().__init__(source, volume)

        self.requester = ctx.author
        self.channel = ctx.channel
        self.data = data

        self.uploader = data.get('uploader')
        self.uploader_url = data.get('uploader_url')
        date = data.get('upload_date')
        self.upload_date = date[6:8] + '.' + date[4:6] + '.' + date[0:4]
        self.title = data.get('title')
        self.thumbnail = data.get('thumbnail')
        self.description = data.get('description')
        self.duration = self.parse_duration(int(data.get('duration')))
        self.tags = data.get('tags')
        self.url = data.get('webpage_url')
        self.views = data.get('view_count')
        self.likes = data.get('like_count')
        self.dislikes = data.get('dislike_count')
        self.stream_url = data.get('url')

    def __str__(self):
        return '**{0.title}** por **{0.uploader}**'.format(self)

    @classmethod
    async def create_source(cls, ctx: commands.Context, search: str, *, loop: asyncio.BaseEventLoop = None):
        loop = loop or asyncio.get_event_loop()

        partial = functools.partial(cls.ytdl.extract_info, search, download=False, process=False)
        data = await loop.run_in_executor(None, partial)

        if data is None:
            raise YTDLError('Couldn\'t find anything that matches `{}`'.format(search))

        if 'entries' not in data:
            process_info = data
        else:
            process_info = None
            for entry in data['entries']:
                if entry:
                    process_info = entry
                    break

            if process_info is None:
                raise YTDLError('Couldn\'t find anything that matches `{}`'.format(search))

        webpage_url = process_info['webpage_url']
        partial = functools.partial(cls.ytdl.extract_info, webpage_url, download=False)
        processed_info = await loop.run_in_executor(None, partial)

        if processed_info is None:
            raise YTDLError('Couldn\'t fetch `{}`'.format(webpage_url))

        if 'entries' not in processed_info:
            info = processed_info
        else:
            info = None
            while info is None:
                try:
                    info = processed_info['entries'].pop(0)
                except IndexError:
                    raise YTDLError('Couldn\'t retrieve any matches for `{}`'.format(webpage_url))

        return cls(ctx, discord.FFmpegPCMAudio(info['url'], **cls.FFMPEG_OPTIONS), data=info)

    @staticmethod
    def parse_duration(duration: int):
        minutes, seconds = divmod(duration, 60)
        hours, minutes = divmod(minutes, 60)
        days, hours = divmod(hours, 24)

        duration = []
        if days > 0:
            duration.append('{} days'.format(days))
        if hours > 0:
            duration.append('{} hours'.format(hours))
        if minutes > 0:
            duration.append('{} minutes'.format(minutes))
        if seconds > 0:
            duration.append('{} seconds'.format(seconds))

        return ', '.join(duration)

class Song:
    __slots__ = ('source', 'requester')

    def __init__(self, source: YTDLSource):
        self.source = source
        self.requester = source.requester

    def create_embed(self):
        embed = (discord.Embed(title='Agora estou cantando',
                               description='```css\n{0.source.title}\n```'.format(self),
                               color=discord.Color.red())
                 .add_field(name='A duração da música é de', value=self.source.duration)
                 .add_field(name='Essa música foi pedida por', value=self.requester.mention)
                 .add_field(name='O autor dessa música é', value='[{0.source.uploader}]({0.source.uploader_url})'.format(self))
                 .add_field(name='URL', value='[Click]({0.source.url})'.format(self))
                 .set_thumbnail(url=self.source.thumbnail))

        return embed

class SongQueue(asyncio.Queue):
    def __getitem__(self, item):
        if isinstance(item, slice):
            return list(itertools.islice(self._queue, item.start, item.stop, item.step))
        else:
            return self._queue[item]

    def __iter__(self):
        return self._queue.__iter__()

    def __len__(self):
        return self.qsize()

    def clear(self):
        self._queue.clear()

    def shuffle(self):
        random.shuffle(self._queue)

    def remove(self, index: int):
        del self._queue[index]

class VoiceState:
    def __init__(self, bot: commands.Bot, ctx: commands.Context):
        self.bot = bot
        self._ctx = ctx

        self.current = None
        self.voice = None
        self.next = asyncio.Event()
        self.songs = SongQueue()

        self._loop = False
        self._volume = 0.5
        self.skip_votes = set()

        self.audio_player = bot.loop.create_task(self.audio_player_task())

    def __del__(self):
        self.audio_player.cancel()

    @property
    def loop(self):
        return self._loop

    @loop.setter
    def loop(self, value: bool):
        self._loop = value

    @property
    def volume(self):
        return self._volume

    @volume.setter
    def volume(self, value: float):
        self._volume = value

    @property
    def is_playing(self):
        return self.voice and self.current

    async def audio_player_task(self):
        while True:
            self.next.clear()

            if not self.loop:
                # Try to get the next song within 15 minutes.
                # If no song will be added to the queue in time,
                # the player will disconnect due to performance
                # reasons.
                try:
                    async with timeout(900):  # 15 minutes
                        self.current = await self.songs.get()
                except asyncio.TimeoutError:
                    self.bot.loop.create_task(self.stop())
                    return

            self.current.source.volume = self._volume
            self.voice.play(self.current.source, after=self.play_next_song)
            await self.current.source.channel.send(embed=self.current.create_embed())

            await self.next.wait()

    def play_next_song(self, error=None):
        if error:
            raise VoiceError(str(error))

        self.next.set()

    def skip(self):
        self.skip_votes.clear()

        if self.is_playing:
            self.voice.stop()

    async def stop(self):
        self.songs.clear()

        if self.voice:
            await self.voice.disconnect()
            self.voice = None

class Music(commands.Cog):

    def __init__(self, bot: commands.Bot):
        self.bot = bot
        self.voice_states = {}

    def get_voice_state(self, ctx: commands.Context):
        state = self.voice_states.get(ctx.guild.id)
        if not state:
            state = VoiceState(self.bot, ctx)
            self.voice_states[ctx.guild.id] = state

        return state

    def cog_unload(self):
        for state in self.voice_states.values():
            self.bot.loop.create_task(state.stop())

    def cog_check(self, ctx: commands.Context):
        if not ctx.guild:
            raise commands.NoPrivateMessage('Esses comandos nas DMs, por favor, utilize os comandos no nosso servidor.')

        return True

    async def cog_before_invoke(self, ctx: commands.Context):
        ctx.voice_state = self.get_voice_state(ctx)

    async def cog_command_error(self, ctx: commands.Context, error: commands.CommandError):
        await ctx.send('Um erro aconteceu, por favor, encaminhar esse código para... mim? : {}'.format(str(error)))

    @commands.command(name='entrar', invoke_without_subcommand=True)
    async def _join(self, ctx: commands.Context):

        destination = ctx.author.voice.channel
        if ctx.voice_state.voice:
            await ctx.voice_state.voice.move_to(destination)
            return

        ctx.voice_state.voice = await destination.connect()

    @commands.command(name='conectar')
    @commands.has_permissions(manage_guild=True)
    async def _summon(self, ctx: commands.Context, *, channel: discord.VoiceChannel = None):

        if not channel and not ctx.author.voice:
            raise VoiceError('Você não especificou nenhum canal de voz para eu conectar.')

        destination = channel or ctx.author.voice.channel
        if ctx.voice_state.voice:
            await ctx.voice_state.voice.move_to(destination)
            return

        ctx.voice_state.voice = await destination.connect()

    @commands.command(name='sair', aliases=['desconectar'])
    @commands.has_permissions(manage_guild=True)
    async def _leave(self, ctx: commands.Context):

        if not ctx.voice_state.voice:
            return await ctx.send('Eu não estou cantando em nenhum canal de voz.')

        await ctx.voice_state.stop()
        del self.voice_states[ctx.guild.id]

    @commands.command(name='volume')
    async def _volume(self, ctx: commands.Context, *, volume: int):

        if not ctx.voice_state.is_playing:
            return await ctx.send('Não estou cantando nada agora')

        if 100 < volume:
            return await ctx.send('O volume precisa estar entre 0 e 100.')

        ctx.voice_state.volume = volume / 100
        await ctx.send('O volume da música foi definido para {}%. Ele será ajustado para a próxima música!'.format(volume))

    @commands.command(name='agora', aliases=['current', 'tocando'])
    async def _now(self, ctx: commands.Context):

        await ctx.send(embed=ctx.voice_state.current.create_embed())

    @commands.command(name='pausar', aliases =["pause"])
    @commands.has_permissions(manage_guild=True)
    async def _pause(self, ctx: commands.Context):

        if not ctx.voice_state.is_playing and ctx.voice_state.voice.is_playing():
            ctx.voice_state.voice.pause()
            await ctx.message.add_reaction('✅')

    @commands.command(name='continuar', aliases=["resume"])
    @commands.has_permissions(manage_guild=True)
    async def _resume(self, ctx: commands.Context):

        if not ctx.voice_state.is_playing and ctx.voice_state.voice.is_paused():
            ctx.voice_state.voice.resume()
            await ctx.message.add_reaction('✅')

    @commands.command(name='reiniciar', aliases=["limpar"])
    @commands.has_permissions(manage_guild=True)
    async def _stop(self, ctx: commands.Context):

        ctx.voice_state.songs.clear()

        if not ctx.voice_state.is_playing:
            ctx.voice_state.voice.stop()
            await ctx.message.add_reaction('✅')

    @commands.command(name='s', aliases=["skip"])
    async def _skip(self, ctx: commands.Context):

        if not ctx.voice_state.is_playing:
            return await ctx.send('Eu não estou cantando nenhuma música atualmente...')

        voter = ctx.message.author
        if voter == ctx.voice_state.current.requester:
            await ctx.message.add_reaction('✅')
            ctx.voice_state.skip()

        if not voter == ctx.voice_state.current.requester:
            await ctx.message.add_reaction('✅')
            ctx.voice_state.skip()

        else:
            await ctx.send('```Pulando essa música, indo para a próxima!```')

    @commands.command(name='fila')
    async def _queue(self, ctx: commands.Context, *, page: int = 1):

        if len(ctx.voice_state.songs) == 0:
            return await ctx.send('Não tem nenhuma música na fila, ainda...')

        items_per_page = 10
        pages = math.ceil(len(ctx.voice_state.songs) / items_per_page)

        start = (page - 1) * items_per_page
        end = start + items_per_page

        queue = ''
        for i, song in enumerate(ctx.voice_state.songs[start:end], start=start):
            queue += '`{0}.` [**{1.source.title}**]({1.source.url})\n'.format(i + 1, song)

        embed = (discord.Embed(description='**{} músicas na fila:**\n\n{}'.format(len(ctx.voice_state.songs), queue))
                 .set_footer(text='Vendo a página {}/{}'.format(page, pages)))
        await ctx.send(embed=embed)

    @commands.command(name='embaralhar')
    async def _shuffle(self, ctx: commands.Context):

        if len(ctx.voice_state.songs) == 0:
            return await ctx.send('```Impossível realizar essa ação: não existem músicas na fila.```')

        ctx.voice_state.songs.shuffle()
        await ctx.message.add_reaction('✅')

    @commands.command(name='remover')
    async def _remove(self, ctx: commands.Context, index: int):

        if len(ctx.voice_state.songs) == 0:
            return await ctx.send('Não quero cantar nenhuma musica ainda')

        ctx.voice_state.songs.remove(index - 1)
        await ctx.message.add_reaction('✅')

    @commands.command(name='loop')
    async def _loop(self, ctx: commands.Context):

        if not ctx.voice_state.is_playing:
            return await ctx.send('Nada está sendo tocado no momento.')

        # Inverse boolean value to loop and unloop.
        ctx.voice_state.loop = not ctx.voice_state.loop
        await ctx.message.add_reaction('✅')

    @commands.command(name='p', aliases=["play"])
    async def _play(self, ctx: commands.Context, *, search: str):

        if not ctx.voice_state.voice:
            await ctx.invoke(self._join)

        async with ctx.typing():
            try:
                source = await YTDLSource.create_source(ctx, search, loop=self.bot.loop)
            except YTDLError as e:
                await ctx.send('Ocorreu um erro enquanto eu processava o seu pedido: {}'.format(str(e)))
            else:
                song = Song(source)

                await ctx.voice_state.songs.put(song)
                await ctx.send('{} foi adicionado na fila'.format(str(source)))

    @_join.before_invoke
    @_play.before_invoke
    async def ensure_voice_state(self, ctx: commands.Context):
        if not ctx.author.voice or not ctx.author.voice.channel:
            raise commands.CommandError('Você não está conectado em nenhum chat de voz!.')

        if ctx.voice_client:
            if ctx.voice_client.channel != ctx.author.voice.channel:
                raise commands.CommandError('Eu já estou cantando em outro chat de voz!.')    

intents = discord.Intents.all()
bot = commands.Bot(command_prefix=['jp '], description='Bot', intents = intents)
bot.add_cog(Music(bot))
       
@bot.command(name="felps")
async def felps(ctx):
    if (ctx.author.voice):
        channel = ctx.message.author.voice.channel
        voice = await channel.connect()
        source = FFmpegPCMAudio('felps.mp3')
        player = voice.play(source)
    
    else:
        await ctx.send("Você precisa estar conectado em um canal de voz para usar esse comando")

#COMANDOS DE UTILIDADE/ADMINISTRAÇÃO ABAIXO :::
@bot.command(name="registrar")
async def register(ctx):
    file = "/Users/JoaoRagazzo/Desktop/bot discord musica/database.xlsx"
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.worksheets[0]
    wsr = wb.active

    channel = bot.get_channel(896511142460014624)
    author = ctx.author
    authorid = ctx.author.id

    if isinstance(ctx.channel, discord.channel.DMChannel):
        pass
    else:
        await ctx.reply("```Por favor, cheque o seu privado!```")
    
    embedVar = discord.Embed(title = "Vincular contas ao Setor III", description="Esse processo pode demorar de 24 a 48 horas.\n\n", color=0xD43030)
    embedVar.add_field(name="Vincular contas", value="Por favor, selecione a plataforma a qual você deseja vincular a sua conta.\n\n`1️⃣ VALORANT`\n`2️⃣ Epic Games`")

    awnser = await author.send(embed=embedVar)

    valid_reaction = ['1️⃣', '2️⃣','✅','❌']

    await awnser.add_reaction('1️⃣')
    await awnser.add_reaction('2️⃣')


    def check(reaction, user):
        return user == ctx.author and str(reaction.emoji) in valid_reaction
    reaction, user = await bot.wait_for('reaction_add', timeout = 60.0, check=check)

    async def rplataforma_(plataforma_):
        plataforma = plataforma_
        n = 0
        for row in wsr.iter_rows(wsr.min_row,wsr.max_row):
            for cell in row:
                if cell.value == author.id and (ws.cell(row=cell.row, column = 3).value) == plataforma_:
                    n = n + 1
                else:
                    break
        
        if n > 0:
            question = await author.send("```Você já tem {} contas(s) registrada(s). Você deseja cadastrar outra conta?```".format(n))
            await question.add_reaction("✅")
            await question.add_reaction("❌")

            def check(reaction,user):
                return user == ctx.author and str(reaction.emoji) in valid_reaction
            reaction, user = await bot.wait_for('reaction_add', timeout = 60.0, check=check)

            if str(reaction.emoji) == "✅":
                pass
            if str(reaction.emoji) == "❌":
                await author.send("```Registro cancelado!```")
                return

        else:
            pass

        await author.send("```Por favor, insira o email/nome de usuário da sua conta do(a) {}:```".format(plataforma))

    if str(reaction.emoji) == '1️⃣':
        plataforma = "VALORANT"
        await rplataforma_(plataforma)

    elif str(reaction.emoji) == '2️⃣':
        plataforma = "Epic Games"
        await rplataforma_(plataforma)

    def check(msg):
        return msg.author == ctx.author

    username = await bot.wait_for("message", timeout = 60, check=check)
    username = username.content

    await author.send("```Por favor, insira sua senha do(a) {}```".format(plataforma))

    password = await bot.wait_for("message", timeout = 60, check=check)
    password = password.content

    embedVar2 = discord.Embed(Title="Vincular conta do {}".format(plataforma), description="Por favor, cheque se suas informações estão corretas:", color=0xD43030)
    embedVar2.add_field(name="**Usuário/Email:**", value="`{}`".format(username), inline=True).add_field(name="**Senha:**", value="`{}`".format(password), inline=True)
    checar = await author.send(embed=embedVar2)

    await checar.add_reaction('✅')
    await checar.add_reaction('❌')

    def check(reaction, user):
        return user == ctx.author and str(reaction.emoji) in valid_reaction
    reaction, user = await bot.wait_for('reaction_add', timeout = 60.0, check=check)

    if str(reaction.emoji) == '✅':

        embedVar3 = discord.Embed(Title="Vinculação de usuário", description="Usuário portador do ID:  `{}`\nUsuário portador do vulgo:  `{}`".format((authorid),(author)), color=0xD43030)
        embedVar3.add_field(name="**Plataforma:**", value="`{}`".format(plataforma), inline = True).add_field(name="\u200b", value="\u200b", inline="True").add_field(name="\u200b", value="\u200b", inline="True")
        embedVar3.add_field(name="**Usuário/Email:**", value="`{}`".format(username), inline=True).add_field(name="**Senha:**", value="`{}`".format(password), inline = True)
        registerfinal = await channel.send(embed=embedVar3)
        await registerfinal.add_reaction('👍🏻')

        await author.send("```Tudo pronto! Seu pedido foi registrado com sucesso!```")
        await checar.delete()

    elif str(reaction.emoji) == '❌':
        await author.send("```Por favor, realize o comando novamente para realizar outro registro```")
        await checar.delete()
        return   

    ws["A{}".format(ws.max_row + 1)] = authorid 
    ws["B{}".format(ws.max_row)] = author.name  
    ws["C{}".format(ws.max_row)] = plataforma
    ws["D{}".format(ws.max_row)] = username
    ws["E{}".format(ws.max_row)] = password
    wb.save(file)

@bot.command(name="roletarussa")
async def rr(ctx):
    membro = ctx.author
    await ctx.author.avatar_url_as(format="png").save(fp="morreu/vitima.png")
    cor = discord.Colour.random()
    n = random.randint(1,6)
    print(n)
    if n == 1:
        vitima = cv2.imread('morreu/vitima.png', cv2.IMREAD_GRAYSCALE)
        (thresh, im_bw) = cv2.threshold(vitima, 128, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)
        cv2.imwrite("morreu/vitima.png", im_bw)
        vitima = cv2.imread("morreu/vitima.png")
        skull = cv2.imread("morreu/skull.jpg")
        skull_resized = cv2.resize(skull, (512, 512))
        img = cv2.resize(vitima, (512, 512))
        vis = cv2.addWeighted(img, 0.2, skull_resized, 0.8, 0.0)
        cv2.imwrite('morreu/morte.png', vis)
        file = discord.File("morreu/morte.png", filename="morte.png")
        time.sleep(0.5)
        morreu = discord.Embed(title="VOCÊ MORREU!!!", description="Você partiu dessa para melhor :(", color = cor)
        morreu.set_image(url="attachment://morte.png")
        await ctx.send(embed=morreu, file=file)
        convite = discord.Embed(title="Não se preocupe!", description="Você morreu, mas te darei uma última chance! Estou encaminhando um convite para você voltar para o nosso servidor", colour = cor)
        await membro.send(embed=convite)
        await membro.send("https://discord.gg/2aYFsYVr6b")
        time.sleep(0.001)
        await membro.kick()
    else:
        time.sleep(0.5)
        sobreviveu = discord.Embed(description="Que sorte. A camara estava vazia...\nVocê teve sorte dessa vez... Mas a morte sabe a hora certa...", colour = cor)
        await ctx.send(embed=sobreviveu)

@bot.command(name="aulas")
async def aulas(ctx):
    today = date.today()
    nowtime = datetime.now()

    t1 = nowtime.strftime("%H:%M:%S")
    t2 = nowtime.strftime("%H")
    s1 = today.isocalendar()[1]
    d1 = today.strftime("%d/%m/%Y")
    d2 = today.weekday()
    d3 = today.strftime("%d")

    if d2 == 6:
        s1 = s1 + 1
    else:
        s1 = s1

    if (s1 % 2) == 0:
        msg1 = "B"
    else:
        msg1 = "A"

    if d2 == 1:
        dds = "terça-feira"
        if msg1 == "A":
            if (int(t2) > 13):
                await ctx.send("Hoje é dia `{}`, sendo uma `{}`. Estamos na semana `{}`. Agora são `{}`".format((d1),(dds),(msg1),(t1)))
                await ctx.send("Estou anexando o horário das aulas de amanhã.", file=discord.File("images/horario_escola/semana_a_quarta.png"))
            else:
                await ctx.send("Hoje é dia `{}`, sendo uma `{}`. Estamos na semana `{}`. Agora são `{}`".format((d1),(dds),(msg1),(t1)))
                await ctx.send("Estou anexando o horário das aulas de hoje.", file=discord.File("images/horario_escola/semana_a_quinta.png"))
        if msg1 == "B":
            if (int(t2) > 13):
                await ctx.send("Hoje é dia `{}`, sendo uma `{}`. Estamos na semana `{}`. Agora são `{}`".format((d1),(dds),(msg1),(t1)))
                await ctx.send("Estou anexando o horário das aulas de amanhã.", file=discord.File("images/horario_escola/semana_b_quarta.png"))
            else:
                await ctx.send("Hoje é dia `{}`, sendo uma `{}`. Estamos na semana `{}`. Agora são `{}`".format((d1),(dds),(msg1),(t1)))
                await ctx.send("Estou anexando o horário das aulas de hoje.", file=discord.File("images/horario_escola/semana_b_quinta.png"))
    elif d2 == 2:
        dds = "quarta-feira"
        if msg1 == "A":
            if (int(t2) > 13):
                await ctx.send("Hoje é dia `{}`, sendo uma `{}`. Estamos na semana `{}`. Agora são `{}`".format((d1),(dds),(msg1),(t1)))
                await ctx.send("Estou anexando o horário das aulas de amanhã.", file=discord.File("semana_a_quinta.png"))
            else:
                await ctx.send("Hoje é dia `{}`, sendo uma `{}`. Estamos na semana `{}`. Agora são `{}`".format((d1),(dds),(msg1),(t1)))
                await ctx.send("Estou anexando o horário das aulas de hoje.", file=discord.File("semana_a_quarta.png"))
        if msg1 == "B":
            if (int(t2) > 13):
                await ctx.send("Hoje é dia `{}`, sendo uma `{}`. Estamos na semana `{}`. Agora são `{}`".format((d1),(dds),(msg1),(t1)))
                await ctx.send("Estou anexando o horário das aulas de amanhã.", file=discord.File("images/horario_escola/semana_b_quinta.png"))
            else:
                await ctx.send("Hoje é dia `{}`, sendo uma `{}`. Estamos na semana `{}`. Agora são `{}`".format((d1),(dds),(msg1),(t1)))
                await ctx.send("Estou anexando o horário das aulas de hoje.", file=discord.File("images/horario_escola/semana_b_quarta.png"))
    elif d2 == 3:
        dds = "quinta-feira"
        if msg1 == "A":
            if (int(t2) > 13):
                await ctx.send("Hoje é dia `{}`, sendo uma `{}`. Estamos na semana `{}`. Agora são `{}`".format((d1),(dds),(msg1),(t1)))
                await ctx.send("Estou anexando o horário das aulas de amanhã.", file=discord.File("images/horario_escola/semana_a_sexta.png"))
            else:
                await ctx.send("Hoje é dia `{}`, sendo uma `{}`. Estamos na semana `{}`. Agora são `{}`".format((d1),(dds),(msg1),(t1)))
                await ctx.send("Estou anexando o horário das aulas de hoje.", file=discord.File("images/horario_escola/semana_a_quinta.png"))
        if msg1 == "B":
            if (int(t2) > 13):
                await ctx.send("Hoje é dia `{}`, sendo uma `{}`. Estamos na semana `{}`. Agora são `{}`".format((d1),(dds),(msg1),(t1)))
                await ctx.send("Estou anexando o horário das aulas de amanhã", file=discord.File("images/horario_escola/semana_b_sexta.png"))
            else:
                await ctx.send("Hoje é dia `{}`, sendo uma `{}`. Estamos na semana `{}`. Agora são `{}`".format((d1),(dds),(msg1),(t1)))
                await ctx.send("Estou anexando o horário das aulas de hoje.", file=discord.File("images/horario_escola/semana_b_quinta.png"))
    elif d2 == 4:
        dds = "sexta-feira"
        if msg1 == "A":
            if (int(t2) > 13):
                await ctx.send("Hoje é dia `{}`, sendo uma `{}`. Estamos na semana `{}`. Agora são `{}`".format((d1),(dds),(msg1),(t1)))
                await ctx.send("Amanhã não tem aula. Bom descanso!")
            else:
                await ctx.send("Hoje é dia `{}`, sendo uma `{}`. Estamos na semana `{}`. Agora são `{}`".format((d1),(dds),(msg1),(t1)))
                await ctx.send("Estou anexando o horário das aulas de hoje.", file=discord.File("images/horario_escola/semana_a_sexta.png"))
        if msg1 == "B":
            if (int(t2) > 13):
                await ctx.send("Hoje é dia `{}`, sendo uma `{}`. Estamos na semana `{}`. Agora são `{}`".format((d1),(dds),(msg1),(t1)))
                await ctx.send("Amanhã não tem aula. Bom descanso!")
            else:
                await ctx.send("Hoje é dia `{}`, sendo uma `{}`. Estamos na semana `{}`. Agora são `{}`".format((d1),(dds),(msg1),(t1)))
                await ctx.send("Estou anexando o horário das aulas de hoje.", file=discord.File("images/horario_escola/semana_b_sexta.png"))
    elif d2 == 5:
        dds = "sábado"
        await ctx.send("Hoje é dia `{}`, sendo uma `{}`. Estamos na semana `{}`. Agora são `{}`".format((d1),(dds),(msg1),(t1)))
        await ctx.send("Amanhã não tem aula. Bom descanso!")
    elif d2 == 6:
        dds = "domingo"
        if msg1 == "A":
            await ctx.send("Hoje é dia `{}`, sendo uma `{}`. Estamos na semana `{}`. Agora são `{}`".format((d1),(dds),(msg1),(t1)))
            await ctx.send("Estou anexando o horário das aulas de amanhã.", file=discord.File("images/horario_escola/semana_a_segunda.png"))
        if msg1 == "B":
            await ctx.send("Hoje é dia `{}`, sendo uma `{}`. Estamos na semana `{}`. Agora são `{}`".format((d1),(dds),(msg1),(t1)))
            await ctx.send("Estou anexando o horário de aulas de amanhã.", file=discord.File("images/horario_escola/semana_b_segunda.png"))
    elif d2 == 0:
        dds = "segunda-feira"
        if msg1 == "A":
            if (int(t2) > 13):
                await ctx.send("Hoje é dia `{}`, sendo uma `{}`. Estamos na semana `{}`. Agora são `{}`".format((d1),(dds),(msg1),(t1)))
                await ctx.send("Estou anexando o horário das aulas de amanhã.", file=discord.File("images/horario_escola/semana_a_terça.png"))
            else:
                await ctx.send("Hoje é dia `{}`, sendo uma `{}`. Estamos na semana `{}`. Agora são `{}`".format((d1),(dds),(msg1),(t1)))
                await ctx.send("Estou anexando o horário das aulas de hoje.", file=discord.File("images/horario_escola/semana_a_segunda.png"))
        if msg1 == "B":
            if (int(t2) > 13):
                await ctx.send("Hoje é dia `{}`, sendo uma `{}`. Estamos na semana `{}`. Agora são `{}`".format((d1),(dds),(msg1),(t1)))
                await ctx.send("Estou anexando o horário das aulas de amanhã.", file=discord.File("images/horario_escola/semana_b_terça.png"))
            else:
                await ctx.send("Hoje é dia `{}`, sendo uma `{}`. Estamos na semana `{}`. Agora são `{}`".format((d1),(dds),(msg1),(t1)))
                await ctx.send("Estou anexando o horário das aulas de hoje.", file=discord.File("images/horario_escola/semana_b_segunda.png"))

    
    if d2 == 5 or d2 == 4 and (int(t2)) > 15:
        return

    if (int(d3) % 2) != 0:
        if (int(t2)) > 15:
            await ctx.send("Amanhã quem vai descer do busão primeiro é o `Neto e o Pedro`")
        else:
            await ctx.send("Hoje quem vai descer do busão primeiro é o `João Paulo`")
    else:
        if (int(t2)) > 15:
            await ctx.send("Amanhã quem vai descer do busão primeiro é o `João Paulo`")
        else:
            await ctx.send("Hoje quem vai descer do busão primeiro é o `Neto e o Pedro`")   

@bot.event
async def on_member_join(member):
    guild = bot.get_guild(482649045412741120)
    channel = bot.get_channel(482654539200200754)
    channel_ = bot.get_channel(902567882540585020)

    member_count = guild.member_count 
    member_count = str(member_count)
    author = member.name
    pic = member.avatar_url
    cor = discord.Colour.random() 

    embedVar = discord.Embed(title="Seja bem-vindo(a)!", description=f"__**{author}**__, seja bem-vindo(a) ao SETOR III\nParabéns! Você é o {member_count}º a entrar no nosso servidor!", colour = cor)
    embedVar.add_field(name="Aproveite o nosso servidor!", value="Aqui temos: \n➝ Uma comunidade em que você pode encontrar membros que possuem os mesmos gostos que você \n➝ Patchnotes exclusivos dos jogos que você acompanha \n➝ Bots exclusivos para você se divertir \n➝ Cargos relacionado com teus gostos e jogos", inline = False)
    embedVar.add_field(name="Pegue seus cargos!", value="Pegue seus cargos em <#887486391699779604> para receber acesso a patchnotes exclusivos dos seus jogos favoritos!")
    embedVar.set_footer(text="✔ SETOR III")
    embedVar.set_image(url="https://i.imgur.com/CEjfz6t.jpg")
    embedVar.set_author(name=author, icon_url=pic)
    embedVar.set_thumbnail(url=pic)

    await channel.send(embed=embedVar)

    await channel_.edit(name=f"Membros: {member_count}")

@bot.event
async def on_member_remove(member):
    guild = bot.get_guild(482649045412741120)
    channel_ = bot.get_channel(902567882540585020)
    member_count = guild.member_count
    member_count = str(member_count)

    await channel_.edit(name=f"Membros: {member_count}")

@bot.event
async def on_message_edit(before, after):
    if before.author.id == 884642846878085200 or before.author.id == 402528814548254720:
        return
    channel = bot.get_channel(893977571077804052)   
    embedVar = discord.Embed(title="__MODIFICAÇÃO DE MENSAGEM__", color=0xD43030)
    embedVar.set_author(name=before.author.name, icon_url=before.author.avatar_url)
    embedVar.add_field(name="**Anterior:**", value="```{}```".format(before.content), inline = True)
    embedVar.add_field(name="**Posterior:**", value = "```{}```".format(after.content), inline = True)

    await channel.send(embed=embedVar)

@bot.command(name="sugestao")
async def sugestao(ctx, *, sugestao: str):
    channel = bot.get_channel(896413999283982346)
    authorid = ctx.author.id
    author = ctx.author
    authorname = ctx.author.name
    icon = ctx.author.avatar_url 

    await ctx.message.reply("```Sugestão enviada com sucesso!```")

    embedVar = discord.Embed(title="Uma sugestão foi enviada:", description="Usuário portador do ID:  `{}`\nUsuário portador do vulgo:  `{}`".format((authorid),(author)))
    embedVar.set_author(name=authorname, icon_url = icon)
    embedVar.add_field(name="Sugestão:", value="{}".format(sugestao), inline=False)
    await channel.send(embed=embedVar)

@bot.command(name="chamar")
async def chamar(ctx, member: discord.Member):
    channel_1 = bot.get_channel(906265009473065011)
    channel_2 = bot.get_channel(906265042335465562)
    member_1 = member.id

    embedVar = discord.Embed(description=f"Chamando {member.mention}\nReaja com \"❌\" para cancelar!", color=discord.Colour.random())
    embedVar.set_author(name=f"{ctx.author.name}", icon_url=f"{ctx.author.avatar_url}")
    embedVar.set_footer(text=f"Solicitado por ⇨ {ctx.author.name}", icon_url=ctx.author.avatar_url)
    alert = await ctx.send(embed = embedVar)
    await alert.add_reaction("❌")
    valid_reaction =["❌"]


    def check(reaction, user):
        return user == ctx.author or member and str(reaction.emoji) in valid_reaction 
    reaction, user = await bot.wait_for('reaction_add', timeout = 60.0, check=check)
    
    try:
        while True:
            await member.move_to(channel_1)
            time.sleep(1)
            await member.move_to(channel_2)
            time.sleep(1)
            if str(reaction.emoji) != None:
                break

    except TimeoutError:
        await ctx.send("ele ta morto")

#COMANDOS PARA TESTE E DESENVOLVIMENTO DO BOT ABAIXO :::
@bot.command(name="aceitar")
async def aceitar(ctx):
    if ctx.author.id == 226539721994665984 or ctx.author.id == 285493211852570634:
        mouse = Controllerm()
        mouse.position = (414, 580)
        time.sleep(0.2)
        mouse.click(Button.left, 1)
        await ctx.reply("Aceitado!")
    else:
        await ctx.reply("Esse comando é exclusivo para desenvolvedores.")
        return

@bot.command(name="cor")
async def color(ctx):
    mouse = Controllerm()
    (x1, y1) = mouse.position

    await ctx.send("Seu mouse está {0}".format(mouse.position))
    await ctx.send("Seu mouse está X: {0} e Y: {1}".format((x1), (y1)))


    px = ImageGrab.grab().load()
    for y in range(0, y1, 1):
        for x in range(0, x1, 1):
            color = px[x, y]

    await ctx.send(color) 

@bot.command(name="registrar2")
async def registrar2(ctx):
    file = "/Users/JoaoRagazzo/Desktop/bot discord musica/database.xlsx"
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.worksheets[0]
    wsr = wb.active
    author = ctx.author
    authorid = ctx.author.id 

    for row in wsr.iter_rows(wsr.min_row,wsr.max_row):
        for cell in row:
            if cell.value == authorid:
                await ctx.reply("Seu ID já está registrado")
                return

    ws["A{}".format(ws.max_row +1)] = authorid
    msg = await ctx.reply("Você é um homem ou uma mulher? Não existe outro genero. Se você não se identifica com um desses generos, clique em **Sair do servidor**")
    await msg.add_reaction("🚹")
    await msg.add_reaction("🚺")

    valid_reaction = ["🚹","🚺"]

    def check(reaction, user):
        return user == ctx.author and str(reaction.emoji) in valid_reaction
    reaction, user = await bot.wait_for('reaction_add', timeout = 60.0, check=check)

    if str(reaction.emoji) == '🚹':
        ws["B{}".format(ws.max_row)] = "Homem"
    elif str(reaction.emoji) == '🚺':
        ws["B{}".format(ws.max_row)] = "Mulher"
    await msg.delete()

    msg1 = await ctx.reply("Em que dia você nasceu?")
    data = await bot.wait_for("message", timeout=60)
    await ctx.send(data.content)
    ws["C{}".format(ws.max_row)] = (data.content)
    await msg1.delete()

    wb.save(file)
    await ctx.reply("Você foi cadastrado!")

@bot.command(name="teste03") #forca
async def teste03(ctx):
    n_ = random.randint(2,4)
    animal = ["animal", "calango", "cachorro", "gato", "girafa"]
    a = animal[0] 
    b = animal[n_]
    n1 = len(b)
    linha = "___ " * n1
    await ctx.send(a + " " + b)
    linhas = await ctx.send("```" + linha + "```")

    def check(msg):
        return msg.author == ctx.author
     
    def findcharacter(string, character):
        pos = []
        for n in range(n1):
            if string[n] == character:
                pos.append(n)
        return pos        

    tentativas = 0
    tentativas_erradas = []
    msg = await ctx.send("vamos ver como voce se sai!")
    cont_tentativa = await ctx.send("tentativas erradas: ")

    while True:
        if tentativas == 6:
            await msg.edit(content="voce morreu")
            await guess.delete()
            break
        else:
            guess = await bot.wait_for("message", timeout = 60, check=check)
            await guess.delete()
        
        guess = guess.content

        if guess in tentativas_erradas:
            await msg.edit(content=f"voce ja tentou essa letra")

        elif findcharacter(b, guess) == []: 
            tentativas += 1
            await msg.edit(content=f"ERROU, tentativas {tentativas}/6")
            tentativas_erradas.append(guess)
            await cont_tentativa.edit(content=" ".join(tentativas_erradas))

        elif findcharacter(b, guess) != []:
            pos = findcharacter(b, guess)
            await msg.edit(content="Acertouuuu")
            line = linha.split(" ")
            print(line)
            print(pos)
            for nn in pos:
                print(nn)
                line = line[nn].replace('___', guess)
                print(line)




@bot.command(name="teste02")
async def teste02(ctx):
    def check(msg):
        return msg.author == ctx.author
    
    wrong_guess = []

    while True:
        teste = await ctx.send(f"suas letras: {wrong_guess}")
        join__ = await bot.wait_for("message", timeout=60, check=check)
        message =join__.content
        wrong_guess.append(message)
        
@bot.command(name="repeat")
async def repeat(ctx):
    def check(msg):
        return msg.author == ctx.author

    await ctx.reply("manda a mensagem")

    message = bot.wait_for("message", timeout=10, check=check)
    await ctx.send(message.content)
    
#CHECAR A LOJA:::
@bot.command(name="epic")
async def epic(ctx):
    erro1 = "Aconteceu algo de errado, por favor, fale com o meu pai para arrumar esse problema."
    awnser2 = await ctx.message.reply("```Estou entrando na sua conta da Epic, esse processo pode demorar um pouco.\n[▯▯▯▯▯▯▯▯▯▯▯▯▯▯▯▯▯▯▯▯] 0%```")

    timeout = time.time() + 60*5
    mouse = Controllerm()
    keyboard = Controllerk()

    mouse.position = (1248, 14)
    time.sleep(0.1)
    mouse.click(Button.left,1)

    if ctx.author.id == 226539721994665984 or ctx.author.id == 285493211852570634:
        username = "joao_ragazzo@hotmail.com"
        password = "36517488Jhon"
    else:
        msg = await ctx.send("```Você não está cadastrado na Epic Games! Fale com o meu pai para ele te cadastrar.```")
        return

    os.startfile("C:/Users/JoaoRagazzo/Desktop/bot discord musica/Epic Games Launcher.lnk")

    while True:
        px = ImageGrab.grab().load()
        for y in range (0, 335, 1):
            for x in range (0, 548, 1):
                color1 = px[x, y]

        if color1 == (50, 97, 163):
            break
        elif time.time() > timeout:
            await ctx.reply(erro1)
            return
        else:
            time.sleep(1)

    mouse.position = (688, 284)
    time.sleep(0.1)
    mouse.click(Button.left, 1)
    time.sleep(0.1)
    await awnser2.edit(content="```Entrando na sua conta...\n[▮▮▮▮▮▮▮▮▮▮▯▯▯▯▯▯▯▯▯▯] 50%```")

    while True:
        px = ImageGrab.grab().load()
        for y in range(0, 517, 1):
            for x in range(0, 577, 1):
                color = px[x, y]

        if color == (16, 74, 130):
            break
        elif time.time() > timeout:
            await ctx.reply(erro1)
        else:
            time.sleep(1)

    mouse.position = (725, 286)
    time.sleep(0.1)
    mouse.click(Button.left, 1)
    time.sleep(0.1)
    keyboard.type(username)
    time.sleep(0.1)
    mouse.position = (696, 367)
    time.sleep(0.1)
    mouse.click(Button.left, 1)
    time.sleep(0.1)
    keyboard.type(password)
    time.sleep(0.1)
    mouse.position = (706, 520)

    while True:
        px = ImageGrab.grab().load()
        for y in range (0, 522, 1):
            for x in range (0, 710, 1):
                color = px[x, y]
        
        if color == (40, 138, 232):
            time.sleep(1)
            mouse.click(Button.left, 1)
            break
        elif time.time() > timeout:
            await ctx.reply(erro1)
            return
        else:
            time.sleep(1)

    await awnser2.edit(content="```Carregando a loja...\n[▮▮▮▮▮▮▮▮▮▮▮▮▮▮▮▮▮▮▯▯] 90%```")

    while True:
        px = ImageGrab.grab().load()
        for y in range (0, 92, 1):
            for x in range (0, 1029, 1):
                color = px[x, y]

        if color == (32, 32, 32):
            break
        elif time.time() > timeout:
            await ctx.reply(erro1)
            return
        else:
            time.sleep(1)

    time.sleep(1)
    mouse.position = (1255, 311)
    time.sleep(0.1)
    mouse.click(Button.left, 15)
    time.sleep(3)

    path = '/Users/JoaoRagazzo/Desktop/bot discord musica/images/prints/epicgames.png' #Caminho onde o print vai ser salvo

    #Coordenadas para o printscreen da loja
    x1 = 377
    y1 = 183
    x2 = 1224
    y2 = 528

    pyautogui.screenshot(path)

    im = Image.open(path)
    im = im.crop((x1, y1, x2, y2))
    im.save(path)
    await awnser2.edit(content="```Só mais um pouquinho...\n[▮▮▮▮▮▮▮▮▮▮▮▮▮▮▮▮▮▮▮▯] 95%```")

    time.sleep(3)
    await awnser2.edit(content="```Tudo feito!\n[▮▮▮▮▮▮▮▮▮▮▮▮▮▮▮▮▮▮▮▮] 100%```")

    awnser = await ctx.message.reply("```Aqui está a loja da Epic Games! Você deseja o jogo gratuito dessa semana?```", file=discord.File('images/prints/epicgames.png'))
    await awnser.add_reaction('✅')
    await awnser.add_reaction('❌')

    valid_reaction = ['✅', '❌']

    def check(reaction, user):
        return user == ctx.author and str(reaction.emoji) in valid_reaction
    reaction, user = await bot.wait_for('reaction_add', timeout = 60.0, check=check)

    if str(reaction.emoji) == '✅':
        await ctx.send("```Essa função ainda está em desenvolvimento, porque meu pai é burro e fez merda. Só vai dar para arrumar semana que vem.```")
 #        awnser3 = ctx.reply("```Um momento...```")
 #        mouse.position = (539, 348)
 #        time.sleep(0.1)
 #        mouse.click(Button.left, 1)
 #        time.sleep(0.1)
 #        await awnser3.edit(content="```Esperando tudo ficar pronto...```")

    elif str(reaction.emoji) == '❌':
        final = await ctx.send("```Entendido!```")

    os.system('taskkill /f /im EpicGamesLauncher.exe')
    os.system('taskkill /f /im EpicOnlineServicesUserHelper.exe')
    await awnser2.delete()
    time.sleep(5)
    await final.delete()

@bot.command(name="loja")
async def loja(ctx):

    #await ctx.send("nao usa isso fdp")
    #return
    global cancelar_
    global loja_ativa_
    cancelar_ = False
    loja_ativa_ = True

    def fechar():
        os.system('taskkill /f /im VALORANT.exe')
        os.system('taskkill /f /im VALORANT-Win64-Shipping.exe')
        os.system('taskkill /f /im RiotClientServices.exe')

    author = ctx.author
    author_id = ctx.author.id
    global author_id_ 
    author_id_ = author_id

    pic = ctx.author.avatar_url
    cor_ = discord.Colour.random()

    resposta = discord.Embed(description="""Estou entrando na sua conta. Esse processo pode demorar. Por favor, aguarde.
                                            Esse processo dura no máximo 5 minutos. 
                                            <a:loading:900811743960367185> [▯▯▯▯▯▯▯▯▯▯▯▯▯▯▯▯▯▯▯▯] 0%""", colour=cor_)
    resposta.set_author(name="Valorant Store", icon_url="https://i.imgur.com/V8P1mTo.png")
    resposta.set_footer(text="Solicitado por ⇨ {}".format(author), icon_url=pic)

    new_01_resposta = discord.Embed(description="Logando na sua conta...\n<a:loading:900811743960367185> [▮▮▮▯▯▯▯▯▯▯▯▯▯▯▯▯▯▯▯▯] 15%", colour=cor_)
    new_01_resposta.set_author(name="Valorant Store", icon_url="https://i.imgur.com/V8P1mTo.png")
    new_01_resposta.set_footer(text="Solicitado por ⇨ {}".format(author), icon_url=pic)

    new_02_resposta = discord.Embed(description="Carregando o Valorant...\n<a:loading:900811743960367185> [▮▮▮▮▮▮▮▮▮▮▯▯▯▯▯▯▯▯▯▯] 50%", colour=cor_)
    new_02_resposta.set_author(name="Valorant Store", icon_url="https://i.imgur.com/V8P1mTo.png")
    new_02_resposta.set_footer(text="Solicitado por ⇨ {}".format(author), icon_url=pic)

    new_03_resposta = discord.Embed(description="Entrando na sua loja... \n<a:loading:900811743960367185> [▮▮▮▮▮▮▮▮▮▮▮▮▮▮▮▮▮▮▯▯] 90%", colour = cor_)
    new_03_resposta.set_author(name="Valorant Store", icon_url="https://i.imgur.com/V8P1mTo.png")
    new_03_resposta.set_footer(text="Solicitado por ⇨ {}".format(author), icon_url=pic)

    new_04_resposta = discord.Embed(description="Só mais um pouquinho... \n<a:loading:900811743960367185> [▮▮▮▮▮▮▮▮▮▮▮▮▮▮▮▮▮▮▮▯] 95%", colour = cor_)
    new_04_resposta.set_author(name="Valorant Store", icon_url="https://i.imgur.com/V8P1mTo.png")
    new_04_resposta.set_footer(text="Solicitado por ⇨ {}".format(author), icon_url=pic)

    new_05_resposta = discord.Embed(description="Tudo pronto! \n[▮▮▮▮▮▮▮▮▮▮▮▮▮▮▮▮▮▮▮▮] 100%", colour = cor_)
    new_05_resposta.set_author(name="Valorant Store", icon_url="https://i.imgur.com/V8P1mTo.png")
    new_05_resposta.set_footer(text="Solicitado por ⇨ {}".format(author), icon_url=pic)

    error = discord.Embed(description="❌ Alguma coisa deu errado :(\nProvavelmente o Valorant atualizou e o código precisa ser atualizado.", colour = cor_)
    error.set_author(name="Valorant Store", icon_url="https://i.imgur.com/V8P1mTo.png")
    error.set_footer(text="Solicitado por ⇨ {}".format(author), icon_url=pic)

    cancelado = discord.Embed(description="❌ Loja cancelada com sucesso!", colour = cor_)
    cancelado.set_author(name="Valorant Store", icon_url="https://i.imgur.com/V8P1mTo.png")
    cancelado.set_footer(text="Solicitado por ⇨ {}".format(author), icon_url=pic)


    timeout = time.time() + 60*5
    mouse = Controllerm()
    keyboard = Controllerk()

    valid_reaction = ['1️⃣', '2️⃣', '3️⃣', '4️⃣', '5️⃣', '6️⃣', '7️⃣', '8️⃣', '9️⃣']

    mouse.position = (1248, 14)
    time.sleep(0.1)
    mouse.click(Button.left,1)

    if ctx.author.id == 226539721994665984 or ctx.author.id == 285493211852570634: #João Paulo 
        username = "Noobinho1337"
        password = "Maggie02092003"
        msgg = await ctx.message.reply(embed = resposta)

    elif ctx.author.id == 549739838518329383: #Gustavo
        username = "ainbootzin"
        password = "G30821564G"
        msgg = await ctx.message.reply(embed = resposta)

    elif ctx.author.id == 286377147663253505: #César
        username = "fallen2547"
        password = "33113515llm"
        msgg = await ctx.message.reply(embed = resposta)
        
    elif ctx.author.id == 392581235802636288: #Neto

        selecionar_conta = discord.Embed(description="Detectamos que você tem mais de uma conta. Qual conta você deseja verificar? \n1️⃣ GodN3to \n2️⃣ Tateno", colour = cor_)
        selecionar_conta.set_author(name="Valorant Store", icon_url="https://i.imgur.com/V8P1mTo.png")
        selecionar_conta.set_footer(text="Solicitado por ⇨ {}".format(author), icon_url=pic)

        msg9 = await ctx.send(embed = selecionar_conta)

        await msg9.add_reaction('1️⃣')
        await msg9.add_reaction('2️⃣')

        def check(reaction, user):
            return user == ctx.author and str(reaction.emoji) in valid_reaction
        reaction, user = await bot.wait_for('reaction_add', timeout = 60.0, check=check) 

        if str(reaction.emoji) == '1️⃣':
            username = "neto908"
            password = "2013.lja.neto"
        
        elif str(reaction.emoji) == '2️⃣':
            username = "Melhorqtateno"
            password = "t4tenoGostosa"

        await msg9.delete()
        msgg = await ctx.message.reply(embed = resposta)

    elif ctx.author.id == 744324004357079102 or ctx.author.id == 397738631747469322: #Pedro

        selecionar_conta = discord.Embed(description="Detectamos que você tem mais de uma conta. Qual conta você deseja verificar? \n1️⃣ proplayer1227 \n2️⃣ DILMA", colour = cor_)
        selecionar_conta.set_author(name="Valorant Store", icon_url="https://i.imgur.com/V8P1mTo.png")
        selecionar_conta.set_footer(text="Solicitado por ⇨ {}".format(author), icon_url=pic)

        msg9 = await ctx.send(embed = selecionar_conta)

        await msg9.add_reaction('1️⃣')
        await msg9.add_reaction('2️⃣')

        def check(reaction, user):
            return user == ctx.author and str(reaction.emoji) in valid_reaction
        reaction, user = await bot.wait_for('reaction_add', timeout = 60.0, check=check) 

        if str(reaction.emoji) == '1️⃣':
            username = "pedrogranola"
            password = "PEDRO@3356"

        if str(reaction.emoji) == '2️⃣':
            username = "pranola"
            password = "Milene@143356"

        await msg9.delete()
        msgg = await ctx.message.reply(embed = resposta)

    elif ctx.author.id == 219933804632997888: #Eduardo
        username = "Unknow021"
        password = "deki210904"
        msgg = await ctx.message.reply(embed = resposta)

    else:
        await ctx.message.reply("Você não está cadastrado. Entre em contato com meu pai para ele te cadastrar!")
        return

    os.startfile('C:/Users/JoaoRagazzo/Desktop/bot discord musica/VALORANT.lnk') #Localização do atalho do Valorant
    
    while True:
        px = ImageGrab.grab().load()
        for y in range (0, 322, 1):
            for x in range (0, 106, 1):
                color = px[x, y]

        if color == (24, 119, 242):
            time.sleep(1)
            break

        elif time.time() > timeout:
            await ctx.reply(embed=error)
            await msgg.delete()
            fechar()
            time.sleep(1)
            return

        elif cancelar_ == True:
            await ctx.reply(embed=cancelado)
            await msgg.delete()
            fechar()
            time.sleep(1)
            return

        else:
            time.sleep(1)
       
    await msgg.edit(embed = new_01_resposta)
 
    mouse.position = (243, 228) #Posição do usuário
    time.sleep(0.1)
    mouse.click(Button.left,1) #Clicar no botão de usuário
    time.sleep(0.1)
    keyboard.type(username) #Atalho do username
    time.sleep(0.1)
    mouse.position = (215, 281) #Posição da senha
    time.sleep(0.1)
    mouse.click(Button.left,1) #Clicar no espaço da senha
    time.sleep(0.1)
    keyboard.type(password) #Atalhao para a senha
    time.sleep(0.1)
    mouse.position = (205, 558) #Posição do botão para entrar
    time.sleep(0.1)
    mouse.click(Button.left,1) #Clicar em logar
    time.sleep(0.1)

    while True: #Cliclar em PLAY se for necessário 
        px = ImageGrab.grab().load()
        for y in range (0, 697, 1):
            for x in range (0, 443, 1):
                color = px[x, y]

        px2 = ImageGrab.grab().load()
        for y2 in range (0, 199, 1):
            for x2 in range (0, 135, 1):
                color2 = px2[x2, y2]

        px3 = ImageGrab.grab().load()
        for y3 in range (0, 187, 1):
            for x3 in range (0, 981, 1):
                color3 = px3[x3, y3]

        if color == (32, 31, 34):
            time.sleep(2)
            mouse.position = (197, 698)  #Posição do botão "PLAY"
            time.sleep(0.1)
            mouse.click(Button.left,1) #Clicar em "PLAY"
            time.sleep(0.1)
            break

        elif color2 == (227, 69, 97):
            break

        elif time.time() > timeout: 
            await ctx.reply(embed=error)
            await msgg.delete()
            fechar()
            time.sleep(1)
            return
        elif cancelar_ == True:
            await ctx.reply(embed=cancelado)
            await msgg.delete()
            fechar()
            time.sleep(1)
            return
        else:
            time.sleep(1)

    while True: #Esperar o VALORANT abrir
        px = ImageGrab.grab().load()
        for y in range (0, 199, 1):
            for x in range (0, 135, 1):
                color = px[x, y]

        if color == (227, 69, 97):
            time.sleep(1)
            break

        elif time.time() > timeout:
            await ctx.reply(embed=error)
            await msgg.delete()
            fechar()
            time.sleep(1)
            return

        elif cancelar_ == True:
            await ctx.reply(embed=cancelado)
            await msgg.delete()
            fechar()
            time.sleep(1)
            return

        else:
            time.sleep(1)   

    await msgg.edit(embed = new_02_resposta)

    while True: #Esperar o VALORANT carregar
        px = ImageGrab.grab().load()
        for y in range (0, 199, 1):
            for x in range (0, 135, 1):
                color = px[x, y]

        if color != (227, 69, 97):
            time.sleep(1)
            break

        elif time.time() > timeout:
            await ctx.reply(embed=error)
            await msgg.delete()
            fechar()
            time.sleep(1)
            return
        elif cancelar_ == True:
            await ctx.reply(embed=cancelado)
            await msgg.delete()
            fechar()
            time.sleep(1)
            return
        else:
            time.sleep(1)

    mouse.position = (1020,20) #Posição de store
    time.sleep(1)
    mouse.click(Button.left,1) #Clicar em store

    await msgg.edit(embed = new_03_resposta)

    time.sleep(3) #Tem que colocar isso se não ele reconhece o "preto" da tela de carregamento no While True abaixo como verdadeiro.

    while True: #Esperar a loja carregar
        px = ImageGrab.grab().load()
        for y in range (0, 619, 1):
            for x in range (0, 988, 1):
                color1 = px[x, y]
        
        px2 = ImageGrab.grab().load()
        for y2 in range (0, 696, 1):
            for x2 in range (0, 627, 1):
                color2 = px2[x2, y2]

        px3 = ImageGrab.grab().load()
        for y3 in range (0, 616, 1):
            for x3 in range (0, 411, 1):
                color3 = px3[x3, y3]
        
        px4 = ImageGrab.grab().load()
        for y4 in range (0, 615, 1):
            for x4 in range (0, 110, 1):
                color4 = px4[x4, y4]

        #px1, px2, px3, px4 são as cores dos intens nas lojas, para ter certeza de que não vai tirar print de um "Very long item name"

        if color1 != (62, 62, 61) and color2 != (62, 62, 61) and color3 != (62, 62, 61) and color4 != (62, 62, 61):
            break
        elif time.time() > timeout:
            await ctx.reply(embed=error)
            await msgg.delete()
            await msg9.delete()
            fechar()
            time.sleep(1)
            return
        elif cancelar_ == True:
            await ctx.reply(embed=cancelado)
            await msgg.delete()

        else:
            time.sleep(1)

    path = '/Users/JoaoRagazzo/Desktop/bot discord musica/images/prints/valorant_store.png' #Localização onde o print vai ficar salvo por tempo indeterminado

    #Coordenadas de onde vai tirar o print 
    x1 = 75
    y1 = 530
    x2 = 1254
    y2 = 710

    pyautogui.screenshot(path)

    im = Image.open(path)
    im = im.crop((x1, y1, x2, y2))
    im.save(path)
    await msgg.edit(embed = new_04_resposta)
    fechar()

    file = discord.File("images/prints/valorant_store.png", filename="valorant_store.png")

    message = discord.Embed(description="Aqui está sua loja:", colour=cor_)
    message.set_author(name="Valorant Store", icon_url="https://i.imgur.com/V8P1mTo.png")
    message.set_footer(text="Solicitado por ⇨ {}".format(author), icon_url=pic)
    message.set_image(url="attachment://valorant_store.png")

    await ctx.message.reply(embed = message, file=file)
    time.sleep(0.1)
    await msgg.edit(embed = new_05_resposta)
    time.sleep(10)
    await msgg.delete()

@bot.command(name="cancelar")
async def cancelar(ctx):
    global cancelar_
    global loja_ativa_
    global author_id_

    try:
        if loja_ativa_ == False:
            await ctx.reply(embed=discord.Embed(description="Não existe um comando de loja ativa no momento!"))
    except NameError:
        await ctx.reply(embed=discord.Embed(description="Não existe um comando de loja ativa no momento!"))

    if author_id_ == ctx.author.id:
        cancelar_ = True
    else:
        await ctx.reply(embed=discord.Embed(description="Somente o autor do comando pode cancelar a loja."))

#COMANDO DE AJUDA PARA COMANDOS DO BOT ABAIXO :::
@bot.command(name="ajuda")
async def ajuda(ctx):
    user = bot.get_user(226539721994665984)
    username = user.name
    discriminator = user.discriminator

    cur_page = 1

    pagina1 = discord.Embed(title="**Unicorninja**", description=f"Prefixo: `jp`", color = 0xD43030)
    pagina1.add_field(name="Descrição", value=f"*Desenvolvido e criado por por* **{username}**.\nCriado em *26 de setembro 2021*\nLinguagem: *Python 3.0*")
    pagina1.add_field(name="Avisos adicionais", value="*O bot ainda está em desenvolvimento. Os comandos estão sujeitos a bugs e lentidão de resposta.*")
    pagina1.add_field(name="Ajuda e comandos", value="Aqui você poderá consultar comandos e ajuda sobre o meu funcionamento.", inline=False)
    pagina1.add_field(name="Atenção!", value="Os parâmetros dos comandos estarão entre chaves.")
    pagina1.set_author(name="Unicorninja", icon_url="https://i.imgur.com/jzHgPG2.png")
    pagina1.set_footer(text=f"Solicitado por ⇨ {ctx.author}  |  Página 1/6", icon_url=f"{ctx.author.avatar_url}")

    pagina2 = discord.Embed(title="**Unicorninja**", description="Comandos de reprodução musical:", color = 0xD43030)
    pagina2.add_field(name="`entrar`", value="Conecta no canal de voz ao qual você está conectado.")
    pagina2.add_field(name="`desconectar`", value="Desconecta do canal de voz ao qual está conectado.")
    pagina2.add_field(name="`volume [0-100]`", value="Estabelece um volume de 0 ~ 100 para as proximas músicas.")
    pagina2.add_field(name="`pausar`", value="Pausa a música atualmente tocando.")
    pagina2.add_field(name="`continuar`", value="Pausa a música anteriormente pausada.")
    pagina2.add_field(name="`fila`", value="Mostra a fila das músicas a serem reproduzidas.")
    pagina2.add_field(name="`embaralhar`", value="Embaralha as músicas a serem reproduzidas.")
    pagina2.add_field(name="`remover [número]`", value="Remove o número correspondente musica da fila")
    pagina2.add_field(name="`play [URL/Nome]`", value="Adiciona uma música a fila de reprodução")
    pagina2.add_field(name="`skip`", value="Pula a música atualmente tocando.")
    pagina2.set_author(name="Unicorninja", icon_url="https://i.imgur.com/jzHgPG2.png")
    pagina2.set_footer(text=f"Solicitado por ⇨ {ctx.author}  |  Página 2/6", icon_url=f"{ctx.author.avatar_url}")

    pagina3 = discord.Embed(title="*Unicorninja*", description="Comandos de utilidades:", color = 0xD43030)
    pagina3.add_field(name="`loja`", value=f"Mostra um print da sua loja temporaria do Valorant\n*Entre em contato com {username} para efetuar o cadastro.*\n*Esse comando pode falhar frequentemente*")
    pagina3.add_field(name="`cancelar`", value="Cancela o comando loja. (É necessário aguardar até a próxima etapa do comando)")
    pagina3.add_field(name="`epic`", value="Mostra um print da loja em promoção da Epic Games.")
    pagina3.add_field(name="`registrar`", value="Envia uma solicitação de cadastro para as plataformas do Valorant ou Epic Games")
    pagina3.add_field(name="`sugestao [Sugestão]`", value="Envia uma sugestão para melhorias no bot")
    pagina3.add_field(name="`aulas`", value="Envia uma imagem contendo as aulas de hoje/amanhã")
    pagina3.set_author(name="Unicorninja", icon_url="https://i.imgur.com/jzHgPG2.png")
    pagina3.set_footer(text=f"Solicitado por ⇨ {ctx.author} | Página 3/6", icon_url=f"{ctx.author.avatar_url}")

    pagina4 = discord.Embed(title="*Unicorninja*", description="Comandos para diversão:", color = 0xD43030)
    pagina4.add_field(name="`roletarussa`", value="Você tem 1 chance em 6 de morrer e ser expulso do servidor")
    pagina4.add_field(name="`felps`", value="O felps entra na call e fala umas coisas... Especiais... E não vai parar...")
    pagina4.set_author(name="Unicorninja", icon_url="https://i.imgur.com/jzHgPG2.png")
    pagina4.set_footer(text=f"Solicitado por ⇨ {ctx.author} | Página 4/6", icon_url=f"{ctx.author.avatar_url}")

    pagina5 = discord.Embed(title="*Unicorninja*", description="Comandos para desenvolvimento e programação:", color=0xD43030)
    pagina5.add_field(name="`cor`", value="Mostra as coordenadas e o valor RGB da cor ao qual o mouse está posicionado na host.")
    pagina5.add_field(name="`aceitar`", value="Aceita a conexão remota na Host via AnyDesk")
    pagina5.add_field(name="`teste[Numero]`", value="Utiizado para desenvolver e testar novas funções sem interferir nas funções principais")
    pagina5.set_author(name="Unicorninja", icon_url="https://i.imgur.com/jzHgPG2.png")
    pagina5.set_footer(text=f"Solicitado por ⇨ {ctx.author} | Página 5/6", icon_url=f"{ctx.author.avatar_url}")

    pagina6 = discord.Embed(title="*Unicorninja*", description = "Informações para contato:", color=0xD43030)
    pagina6.add_field(name="Email:", value="joao_ragazzo@hotmail.com")
    pagina6.add_field(name="Discord:", value=f"{username}#{discriminator}")
    pagina6.set_author(name="Unicorninja", icon_url="https://i.imgur.com/jzHgPG2.png")
    pagina6.set_footer(text=f"Solicitado por ⇨ {ctx.author} | Página 6/6", icon_url=f"{ctx.author.avatar_url}")

    message = await ctx.reply(embed=pagina1)

    valid_reaction = ['◀️','▶️']

    await message.add_reaction("◀️")
    await message.add_reaction("▶️")

    def check(reaction, user):
        return user == ctx.author and str(reaction.emoji) in valid_reaction

    while True:
        try:
            reaction, user = await bot.wait_for('reaction_add', timeout = 90.0, check=check)
            if str(reaction.emoji) == "◀️" and cur_page == 1: #Página mínima 
                await message.remove_reaction(reaction, user)

            if str(reaction.emoji) == "▶️" and cur_page == 6: #Página máxima
                await message.remove_reaction(reaction, user)

            elif str(reaction.emoji) == "▶️" and cur_page != 6:
                cur_page += 1
                if cur_page == 2:
                    await message.edit(embed=pagina2)
                    await message.remove_reaction(reaction, user)
                if cur_page == 3:
                    await message.edit(embed=pagina3)
                    await message.remove_reaction(reaction, user)
                if cur_page == 4:
                    await message.edit(embed=pagina4)
                    await message.remove_reaction(reaction, user)
                if cur_page == 5:
                    await message.edit(embed=pagina5)
                    await message.remove_reaction(reaction, user)
                if cur_page ==6:
                    await message.edit(embed=pagina6)
                    await message.remove_reaction(reaction, user)

            elif str(reaction.emoji) == "◀️":
                cur_page -= 1
                if cur_page == 1:
                    await message.edit(embed=pagina1)
                    await message.remove_reaction(reaction, user)
                if cur_page == 2:
                    await message.edit(embed=pagina2)
                    await message.remove_reaction(reaction, user)
                if cur_page == 3:
                    await message.edit(embed=pagina3)
                    await message.remove_reaction(reaction, user)
                if cur_page == 4:
                    await message.edit(embed=pagina4)
                    await message.remove_reaction(reaction, user)
                if cur_page == 5:
                    await message.edit(embed=pagina5)
                    await message.remove_reaction(reaction, user)

        except asyncio.TimeoutError:
            await message.delete()
            break

#CHECAR EVENTOS PARA DAR CARGOS EM RELAÇÃO A UMA MENSAGEM ABAIXO :::
@bot.event
async def on_raw_reaction_add(payload):

    ourMessageID = 887498004884516864
    if ourMessageID == payload.message_id:
        member = payload.member
        guild = member.guild
        emoji = payload.emoji.name
        if emoji == '0️⃣':
            role = discord.utils.get(guild.roles, name="🌎 Fortnite")
        elif emoji == '1️⃣':
            role = discord.utils.get(guild.roles, name="🌎 Tom Clancy's Rainbow Six")
        elif emoji == '2️⃣':
            role = discord.utils.get(guild.roles, name="🌎 League Of Legends")
        elif emoji == '3️⃣':
            role = discord.utils.get(guild.roles, name="🌎 Counter Strike:Global Offensive")
        elif emoji == '4️⃣':
            role = discord.utils.get(guild.roles, name="🌎 Players Unknow Battlegrounds")
        elif emoji == '5️⃣':
            role = discord.utils.get(guild.roles, name="🌎 Overwatch")
        elif emoji == '6️⃣':
            role = discord.utils.get(guild.roles, name="🌎 Rocket League")
        elif emoji == '7️⃣':
            role = discord.utils.get(guild.roles, name="🌎 Minecraft")
        elif emoji == '8️⃣':
            role = discord.utils.get(guild.roles, name="🌎 World of Warcraft")
        elif emoji == '9️⃣':
            role = discord.utils.get(guild.roles, name="🌎 Rust")
        elif emoji == '🔟':
            role = discord.utils.get(guild.roles, name="🌎 Dota 2")
        elif emoji == '🟫':
            role = discord.utils.get(guild.roles, name="🌎 Factorio")
        elif emoji == '🟥':
            role = discord.utils.get(guild.roles, name="🌎 Heartstone")
        elif emoji == '🟧':
            role = discord.utils.get(guild.roles, name="🌎 Diablo III")
        elif emoji == '🟨':
            role = discord.utils.get(guild.roles, name="🌎 Team Fortress 2")
        elif emoji == '🟩':
            role = discord.utils.get(guild.roles, name="🌎 Grand Theft Auto V")
        elif emoji == '🟦':
            role = discord.utils.get(guild.roles, name="🌎 Payday 2")
        elif emoji == '🟪':
            role = discord.utils.get(guild.roles, name="🌎 Sea of Thieves")
        elif emoji == '⬛️':
            role = discord.utils.get(guild.roles, name="🌎 Garry's Mod")
        elif emoji == '⬜️':
            role = discord.utils.get(guild.roles, name="🌎 Smite")    
        await member.add_roles(role) 

    ourMessageID1 = 887519438881062992
    if ourMessageID1 == payload.message_id:
        member = payload.member
        guild = member.guild
        emoji = payload.emoji.name
        if emoji == '0️⃣':
            role = discord.utils.get(guild.roles, name="🎶 Clássica")
        elif emoji == '1️⃣':
            role = discord.utils.get(guild.roles, name="🎶 Funk")
        elif emoji == '2️⃣':
            role = discord.utils.get(guild.roles, name="🎶 Sertanjo")
        elif emoji == '3️⃣':
            role = discord.utils.get(guild.roles, name="🎶 Rap")
        elif emoji == '4️⃣':
            role = discord.utils.get(guild.roles, name="🎶 Trap")
        elif emoji == '5️⃣':
            role = discord.utils.get(guild.roles, name="🎶 Lo-fi")
        elif emoji == '6️⃣':
            role = discord.utils.get(guild.roles, name="🎶 Eletro")
        elif emoji == '7️⃣':
            role = discord.utils.get(guild.roles, name="🎶 Forró")
        elif emoji == '8️⃣':
            role = discord.utils.get(guild.roles, name="🎶 Pagode")
        elif emoji == '9️⃣':
            role = discord.utils.get(guild.roles, name="🎶 Samba")
        elif emoji == '🔟':
            role = discord.utils.get(guild.roles, name="🎶 Axé")
        elif emoji == '🟫':
            role = discord.utils.get(guild.roles, name="🎶 MPB")
        elif emoji == '🟥':
            role = discord.utils.get(guild.roles, name="🎶 Hip-Hop")
        elif emoji == '🟧':
            role = discord.utils.get(guild.roles, name="🎶 Pop")
        elif emoji == '🟨':
            role = discord.utils.get(guild.roles, name="🎶 Rock")
        elif emoji == '🟩':
            role = discord.utils.get(guild.roles, name="🎶 Sad Music")
        elif emoji == '🟦':
            role = discord.utils.get(guild.roles, name="🎶 Metal") 
        await member.add_roles(role)
    
    ourMessageID2 = 887525283366830110
    if ourMessageID2 == payload.message_id:
        member = payload.member
        guild = member.guild
        emoji = payload.emoji.name
        if emoji == '0️⃣':
            role = discord.utils.get(guild.roles, name="💳 Norte")
        elif emoji == '1️⃣':
            role = discord.utils.get(guild.roles, name="💳 Nordeste")
        elif emoji == '2️⃣':
            role = discord.utils.get(guild.roles, name="💳 Sudeste")
        elif emoji == '3️⃣':
            role = discord.utils.get(guild.roles, name="💳 Sul")
        elif emoji == '4️⃣':
            role = discord.utils.get(guild.roles, name="💳 Centro-Oeste")
        elif emoji =='💻':
            role = discord.utils.get(guild.roles, name="💻 Computador")
        elif emoji == '🎮':
            role = discord.utils.get(guild.roles, name="🎮 Console")
        elif emoji == '📱':
            role = discord.utils.get(guild.roles, name="📱  Celular")
        elif emoji == '💞':
            role = discord.utils.get(guild.roles, name="💞 Namorando")
        elif emoji == '💕':
            role = discord.utils.get(guild.roles, name="💕 Enrolado")
        elif emoji == '👨‍💼':
            role = discord.utils.get(guild.roles, name="👨‍💼 Solteiro")
        elif emoji == '👩‍💼':
            role = discord.utils.get(guild.roles, name="👩‍💼 Solteira")
        await member.add_roles(role)

@bot.event
async def on_raw_reaction_remove(payload):
    ourMessageID = 887498004884516864

    if ourMessageID == payload.message_id:
        guild = await(bot.fetch_guild(payload.guild_id))
        emoji = payload.emoji.name
        if emoji == '0️⃣':
            role = discord.utils.get(guild.roles, name="🌎 Fortnite")
        elif emoji == '1️⃣':
            role = discord.utils.get(guild.roles, name="🌎 Tom Clancy's Rainbow Six")
        elif emoji == '2️⃣':
            role = discord.utils.get(guild.roles, name="🌎 League Of Legends")
        elif emoji == '3️⃣':
            role = discord.utils.get(guild.roles, name="🌎 Counter Strike:Global Offensive")
        elif emoji == '4️⃣':
            role = discord.utils.get(guild.roles, name="🌎 Players Unknow Battlegrounds")
        elif emoji == '5️⃣':
            role = discord.utils.get(guild.roles, name="🌎 Overwatch")
        elif emoji == '6️⃣':
            role = discord.utils.get(guild.roles, name="🌎 Rocket League")
        elif emoji == '7️⃣':
            role = discord.utils.get(guild.roles, name="🌎 Minecraft")
        elif emoji == '8️⃣':
            role = discord.utils.get(guild.roles, name="🌎 World of Warcraft")
        elif emoji == '9️⃣':
            role = discord.utils.get(guild.roles, name="🌎 Rust")
        elif emoji == '🔟':
            role = discord.utils.get(guild.roles, name="🌎 Dota 2")
        elif emoji == '🟫':
            role = discord.utils.get(guild.roles, name="🌎 Factorio")
        elif emoji == '🟥':
            role = discord.utils.get(guild.roles, name="🌎 Heartstone")
        elif emoji == '🟧':
            role = discord.utils.get(guild.roles, name="🌎 Diablo III")
        elif emoji == '🟨':
            role = discord.utils.get(guild.roles, name="🌎 Team Fortress 2")
        elif emoji == '🟩':
            role = discord.utils.get(guild.roles, name="🌎 Grand Theft Auto V")
        elif emoji == '🟦':
            role = discord.utils.get(guild.roles, name="🌎 Payday 2")
        elif emoji == '🟪':
            role = discord.utils.get(guild.roles, name="🌎 Sea of Thieves")
        elif emoji == '⬛️':
            role = discord.utils.get(guild.roles, name="🌎 Garry's Mod")
        elif emoji == '⬜️':
            role = discord.utils.get(guild.roles, name="🌎 Smite")        
        member = await(guild.fetch_member(payload.user_id))
        await member.remove_roles(role)

    ourMessageID1 = 887519438881062992
    if ourMessageID1 == payload.message_id:
        guild = await(bot.fetch_guild(payload.guild_id))
        emoji = payload.emoji.name
        if emoji == '0️⃣':
            role = discord.utils.get(guild.roles, name="🎶 Clássica")
        elif emoji == '1️⃣':
            role = discord.utils.get(guild.roles, name="🎶 Funk")
        elif emoji == '2️⃣':
            role = discord.utils.get(guild.roles, name="🎶 Sertanjo")
        elif emoji == '3️⃣':
            role = discord.utils.get(guild.roles, name="🎶 Rap")
        elif emoji == '4️⃣':
            role = discord.utils.get(guild.roles, name="🎶 Trap")
        elif emoji == '5️⃣':
            role = discord.utils.get(guild.roles, name="🎶 Lo-fi")
        elif emoji == '6️⃣':
            role = discord.utils.get(guild.roles, name="🎶 Eletro")
        elif emoji == '7️⃣':
            role = discord.utils.get(guild.roles, name="🎶 Forró")
        elif emoji == '8️⃣':
            role = discord.utils.get(guild.roles, name="🎶 Pagode")
        elif emoji == '9️⃣':
            role = discord.utils.get(guild.roles, name="🎶 Samba")
        elif emoji == '🔟':
            role = discord.utils.get(guild.roles, name="🎶 Axé")
        elif emoji == '🟫':
            role = discord.utils.get(guild.roles, name="🎶 MPB")
        elif emoji == '🟥':
            role = discord.utils.get(guild.roles, name="🎶 Hip-Hop")
        elif emoji == '🟧':
            role = discord.utils.get(guild.roles, name="🎶 Pop")
        elif emoji == '🟨':
            role = discord.utils.get(guild.roles, name="🎶 Rock")
        elif emoji == '🟩':
            role = discord.utils.get(guild.roles, name="🎶 Sad Music")
        elif emoji == '🟦':
            role = discord.utils.get(guild.roles, name="🎶 Metal") 
        member = await(guild.fetch_member(payload.user_id))
        await member.remove_roles(role)

    ourMessageID2 = 887525283366830110
    if ourMessageID2 == payload.message_id:
        guild = await(bot.fetch_guild(payload.guild_id))
        emoji = payload.emoji.name
        if emoji == '0️⃣':
            role = discord.utils.get(guild.roles, name="💳 Norte")
        elif emoji == '1️⃣':
            role = discord.utils.get(guild.roles, name="💳 Nordeste")
        elif emoji == '2️⃣':
            role = discord.utils.get(guild.roles, name="💳 Sudeste")
        elif emoji == '3️⃣':
            role = discord.utils.get(guild.roles, name="💳 Sul")
        elif emoji == '4️⃣':
            role = discord.utils.get(guild.roles, name="💳 Centro-Oeste")
        elif emoji =='💻':
            role = discord.utils.get(guild.roles, name="💻 Computador")
        elif emoji == '🎮':
            role = discord.utils.get(guild.roles, name="🎮 Console")
        elif emoji == '📱':
            role = discord.utils.get(guild.roles, name="📱  Celular")
        elif emoji == '💞':
            role = discord.utils.get(guild.roles, name="💞 Namorando")
        elif emoji == '💕':
            role = discord.utils.get(guild.roles, name="💕 Enrolado")
        elif emoji == '👨‍💼':
            role = discord.utils.get(guild.roles, name="👨‍💼 Solteiro")
        elif emoji == '👩‍💼':
            role = discord.utils.get(guild.roles, name="👩‍💼 Solteira")
        member = await(guild.fetch_member(payload.user_id))
        if member is not None:
            await member.remove_roles(role)
        else:
            print("Um ser não identificado tentou remover um cargo não identificado. Que bizarro.")

@bot.event
async def on_ready():  
    activity = discord.Game(name="e pensando nela 😔", type = 3)
    await bot.change_presence(status=discord.Status.online, activity=activity)
    print('<=========================================>')
    print('Bot online com sucesso:\n{0.user.name}\n{0.user.id}'.format(bot))
    print('Aguardando comandos e a disposição do usuário')
    print('<=========================================>')

bot.run('ODg0NjQyODQ2ODc4MDg1MjAw.YTbd2g.pyKsrTOiYoL-q8SXg-gv2odzx_E')